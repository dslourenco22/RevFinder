"""Excel report generation for RevFinder comparisons."""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .engine import DiffResult, discrepancy_log, filter_changes


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
STATUS_FILLS = {
    "added": PatternFill("solid", fgColor="D9EAD3"),
    "removed": PatternFill("solid", fgColor="F4CCCC"),
    "modified": PatternFill("solid", fgColor="FFF2CC"),
    "unchanged": PatternFill("solid", fgColor="E7E6E6"),
}


def build_excel_report(diff: DiffResult) -> BytesIO:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _summary_dataframe(diff).to_excel(writer, index=False, sheet_name="Summary")
        diff.document_changes.to_excel(writer, index=False, sheet_name="Document Changes")
        diff.comparison.to_excel(writer, index=False, sheet_name="All Changes")
        filter_changes(diff, "added").to_excel(writer, index=False, sheet_name="Added")
        filter_changes(diff, "removed").to_excel(writer, index=False, sheet_name="Removed")
        filter_changes(diff, "modified").to_excel(writer, index=False, sheet_name="Modified")
        filter_changes(diff, "unchanged").to_excel(writer, index=False, sheet_name="Unchanged")
        discrepancy_log(diff).to_excel(writer, index=False, sheet_name="Discrepancy Log")
        diff.old_items.to_excel(writer, index=False, sheet_name="Raw Old")
        diff.new_items.to_excel(writer, index=False, sheet_name="Raw New")

        for worksheet in writer.book.worksheets:
            _style_sheet(worksheet)
            if worksheet.title in {"All Changes", "Added", "Removed", "Modified", "Unchanged"}:
                _style_status_rows(worksheet)

    output.seek(0)
    return output


def _summary_dataframe(diff: DiffResult) -> pd.DataFrame:
    rows = [
        ("Old source", diff.old_document.source_name),
        ("New source", diff.new_document.source_name),
        ("Old parser", diff.old_document.parser),
        ("New parser", diff.new_document.parser),
        ("Old document type", diff.old_document.document_type),
        ("New document type", diff.new_document.document_type),
        ("Old line items", diff.summary["old_items"]),
        ("New line items", diff.summary["new_items"]),
        ("Added", diff.summary["added"]),
        ("Removed", diff.summary["removed"]),
        ("Modified", diff.summary["modified"]),
        ("Unchanged", diff.summary["unchanged"]),
        ("Document changes", diff.summary.get("document_changes", 0)),
    ]
    return pd.DataFrame(rows, columns=["Metric", "Value"])


def _style_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in worksheet.columns:
        values = [str(cell.value or "") for cell in column_cells]
        width = min(max(len(value) for value in values) + 2, 64)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max(width, 12)

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.auto_filter.ref = worksheet.dimensions


def _style_status_rows(worksheet) -> None:
    headers = [str(cell.value or "") for cell in worksheet[1]]
    if "status" not in headers:
        return
    status_col = headers.index("status") + 1
    for row in worksheet.iter_rows(min_row=2):
        status = str(row[status_col - 1].value or "").lower()
        fill = STATUS_FILLS.get(status)
        if fill:
            for cell in row:
                cell.fill = fill
